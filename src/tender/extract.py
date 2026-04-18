"""
Stage: extract – text extraction + metadata sidecar generation.

Supported formats:
  .pdf   -> pdfplumber (primary) / pypdf fallback; tables rendered inline
  .docx  -> python-docx (paragraphs + tables)
  .xlsx  -> openpyxl (sheet-by-sheet tabular text)
  ~$     -> Office temp/lock files skipped entirely
"""
import json
import re
from pathlib import Path

from .types import DocumentRecord

try:
    import pdfplumber          # type: ignore
    _PDFPLUMBER = True
except ImportError:
    _PDFPLUMBER = False

try:
    from pypdf import PdfReader  # type: ignore
    _PYPDF = True
except ImportError:
    _PYPDF = False

try:
    import docx as _docx_lib   # type: ignore  (python-docx)
    _DOCX = True
except ImportError:
    _DOCX = False

try:
    import openpyxl            # type: ignore
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

def _extract_pdf_pdfplumber(path: str) -> tuple[str, list[dict]]:
    pages_meta: list[dict] = []
    parts: list[str] = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
            table_texts: list[str] = []
            for table in page.extract_tables():
                rows = [" | ".join(str(c or "").strip() for c in row) for row in table]
                table_texts.append("\n".join(rows))
            page_content = text + ("\n\n[TABLES]\n" + "\n\n".join(table_texts) if table_texts else "")
            parts.append(page_content)
            pages_meta.append({
                "page": i,
                "char_count": len(page_content),
                "has_tables": bool(table_texts),
                "width": float(page.width),
                "height": float(page.height),
            })
    return "\n\n--- PAGE BREAK ---\n\n".join(parts), pages_meta


def _extract_pdf_pypdf(path: str) -> tuple[str, list[dict]]:
    pages_meta: list[dict] = []
    parts: list[str] = []
    reader = PdfReader(path)
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        parts.append(text)
        pages_meta.append({"page": i, "char_count": len(text), "has_tables": False})
    return "\n\n--- PAGE BREAK ---\n\n".join(parts), pages_meta


def _extract_pdf(path: str) -> tuple[str, list[dict], str]:
    if _PDFPLUMBER:
        try:
            text, pages = _extract_pdf_pdfplumber(path)
            return text, pages, "pdfplumber"
        except Exception:
            pass
    if _PYPDF:
        try:
            text, pages = _extract_pdf_pypdf(path)
            return text, pages, "pypdf"
        except Exception:
            pass
    return "", [], "failed"


# ---------------------------------------------------------------------------
# DOCX helper
# ---------------------------------------------------------------------------

def _extract_docx(path: str) -> tuple[str, str]:
    """Render paragraphs + tables from a .docx as plain text."""
    doc = _docx_lib.Document(path)
    parts: list[str] = []
    for block in doc.element.body:
        tag = block.tag.split("}")[-1] if "}" in block.tag else block.tag
        if tag == "p":
            para_obj = _docx_lib.text.paragraph.Paragraph(block, doc)
            text = para_obj.text.strip()
            if text:
                parts.append(text)
        elif tag == "tbl":
            from docx.table import Table  # type: ignore
            tbl = Table(block, doc)
            rows: list[str] = []
            for row in tbl.rows:
                cells = [cell.text.strip().replace("\n", " ") for cell in row.cells]
                rows.append(" | ".join(cells))
            if rows:
                parts.append("[TABLE]\n" + "\n".join(rows))
    return "\n\n".join(parts), "python-docx"


# ---------------------------------------------------------------------------
# XLSX helper
# ---------------------------------------------------------------------------

def _extract_xlsx(path: str) -> tuple[str, str]:
    """Render each worksheet as a named pipe-delimited table."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"[SHEET: {sheet_name}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts), "openpyxl"


# ---------------------------------------------------------------------------
# Misc helpers
# ---------------------------------------------------------------------------

def _amendment_number(filename: str) -> int | None:
    m = re.search(r"amendment\+?[\s_-]*0*([0-9]+)", filename, flags=re.IGNORECASE)
    return int(m.group(1)) if m else None


def _is_temp_file(filename: str) -> bool:
    return filename.startswith("~$")


# ---------------------------------------------------------------------------
# Content-based classification
# ---------------------------------------------------------------------------

# Patterns scanned against the first ~2 000 chars of extracted text.
# Each entry: (regex, doc_type)  — first match wins.
_CONTENT_RULES: list[tuple[str, str]] = [
    (r"drawing\s+index|sheet\s+index|revision\s+history.*dwg|drawing\s+list", "drawings"),
    (r"(?:section|div(?:ision)?)\s+0?1\s+\d{2}\s+\d{2}|project\s+manual|table\s+of\s+contents.*division", "solicitation"),
    (r"statement\s+of\s+work|scope\s+of\s+work|performance\s+work\s+statement", "solicitation"),
    (r"wage\s+determination|davis.bacon|service\s+contract\s+act", "wage_determination"),
    (r"amendment\s+(?:no\.?|number)?\s*\d+|this\s+amendment\s+modifies", "amendment"),
    (r"solicitation\s+(?:no\.?|number)|request\s+for\s+(?:proposal|quotation|bid)", "solicitation"),
    (r"past\s+performance\s+questionnaire|ppq", "past_performance"),
    (r"subcontracting\s+plan", "subcontracting_plan"),
    (r"infection\s+control\s+risk|pre.?construction\s+risk\s+assessment", "admin"),
    (r"site\s+visit\s+(?:sign|instructions)|attendance\s+sheet", "admin"),
    (r"questions?\s+and\s+answers|technical\s+questions?\s+(?:and\s+)?responses?", "admin"),
]

# RAG strategy by doc_type
_RAG_STRATEGY: dict[str, str] = {
    "drawings":            "reference_only",   # huge, coordinate-heavy, not narrative
    "unknown":             "reference_only",   # unclassified — don't surface by default
    "solicitation":        "full",
    "amendment":           "full",
    "sf1442":              "full",
    "pricing":             "full",
    "contract":            "full",
    "wage_determination":  "full",
    "past_performance":    "full",
    "bonding":             "full",
    "subcontracting_plan": "full",
    "admin":               "full",
    "tender_summary":      "full",
}


def _content_classify(text: str, pages_meta: list[dict]) -> str | None:
    """
    Try to infer doc_type from extracted text and page dimensions.
    Returns a doc_type string, or None if no confident match.
    """
    # Landscape-majority heuristic: drawing sets are almost always wider than tall
    if pages_meta:
        landscape_count = sum(
            1 for p in pages_meta if p.get("width", 0) > p.get("height", 0)
        )
        if landscape_count / len(pages_meta) > 0.6:
            return "drawings"

    # Text pattern scan on first 2 000 chars
    snippet = text[:2000].lower()
    for pattern, doc_type in _CONTENT_RULES:
        if re.search(pattern, snippet):
            return doc_type

    return None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def extract_all(
    records: list[DocumentRecord],
    out_dir: str,
    solicitation_id: str,
) -> list[DocumentRecord]:
    """Extract text and write .txt + .meta.json sidecars for every record."""
    extract_root = Path(out_dir) / solicitation_id / "extracted"
    extract_root.mkdir(parents=True, exist_ok=True)

    valid_nos = [
        n for n in (_amendment_number(r.filename) for r in records if r.doc_type == "amendment")
        if n is not None
    ]
    max_amendment = max(valid_nos) if valid_nos else None

    kept: list[DocumentRecord] = []
    for rec in records:
        if _is_temp_file(rec.filename):
            continue
        kept.append(rec)

        section_dir = extract_root / rec.section
        section_dir.mkdir(parents=True, exist_ok=True)
        stem = Path(rec.filename).stem
        amend_no = _amendment_number(rec.filename) if rec.doc_type == "amendment" else None

        meta: dict = {
            "solicitation_id":        solicitation_id,
            "filename":               rec.filename,
            "rel_path":               rec.rel_path,
            "sha256":                 rec.sha256,
            "size_bytes":             rec.size_bytes,
            "doc_type":               rec.doc_type,
            "section":                rec.section,
            "sort_key":               rec.sort_key,
            "amendment_no":           amend_no,
            "is_latest_amendment":    amend_no is not None and amend_no == max_amendment,
            "extraction_method":      None,
            "total_pages":            None,
            "total_chars":            None,
            "has_extractable_text":   False,
            "pages":                  [],
            # ── new classification fields ──────────────────────────────────
            "classification_source":  rec.classification_source,   # filename | content | unmatched
            "content_classification": None,   # filled in after extraction
            "rag_strategy":           None,   # filled in after extraction
            # ──────────────────────────────────────────────────────────────
            "dify_doc_name":          f"[{rec.doc_type.upper()}] {rec.filename}",
            "dify_tags":              [rec.doc_type, solicitation_id, rec.section],
        }

        text = ""

        if rec.ext == ".pdf":
            source_path = rec.rendered_pdf if rec.rendered_pdf else rec.abs_path
            text, pages_meta, method = _extract_pdf(source_path)
            meta["extraction_method"]    = method
            meta["total_pages"]          = len(pages_meta)
            meta["total_chars"]          = len(text.strip())
            meta["has_extractable_text"] = len(text.strip()) > 50
            meta["pages"]                = pages_meta

        elif rec.ext == ".docx" and _DOCX:
            try:
                text, method = _extract_docx(rec.abs_path)
                meta["extraction_method"]    = method
                meta["total_chars"]          = len(text.strip())
                meta["has_extractable_text"] = len(text.strip()) > 50
            except Exception as exc:
                meta["extraction_method"] = f"failed: {exc}"

        elif rec.ext == ".xlsx" and _OPENPYXL:
            try:
                text, method = _extract_xlsx(rec.abs_path)
                meta["extraction_method"]    = method
                meta["total_chars"]          = len(text.strip())
                meta["has_extractable_text"] = len(text.strip()) > 50
            except Exception as exc:
                meta["extraction_method"] = f"failed: {exc}"

        else:
            meta["extraction_method"] = "unsupported_format"

        # ── Content-based classification override ─────────────────────────
        if rec.ext == ".pdf" and meta["has_extractable_text"]:
            content_type = _content_classify(text, meta["pages"])
            meta["content_classification"] = content_type

            if content_type and content_type != meta["doc_type"]:
                # Content signal wins over an unmatched filename; also
                # always trust the landscape/drawings signal regardless.
                if meta["classification_source"] == "unmatched" or content_type == "drawings":
                    meta["doc_type"] = content_type
                    meta["classification_source"] = "content"
                    # Re-derive section for drawings
                    if content_type == "drawings":
                        meta["section"] = "11_Drawings"
                    # Update dify fields to reflect new doc_type
                    meta["dify_doc_name"] = f"[{content_type.upper()}] {rec.filename}"
                    meta["dify_tags"][0] = content_type
            elif content_type:
                # Content agreed with filename classification
                meta["content_classification"] = content_type

        # ── RAG strategy ──────────────────────────────────────────────────
        meta["rag_strategy"] = _RAG_STRATEGY.get(meta["doc_type"], "full")

        (section_dir / f"{stem}.txt").write_text(text, encoding="utf-8")
        (section_dir / f"{stem}.meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")

    records[:] = kept
    return records
